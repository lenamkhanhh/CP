import unittest
from collections import Counter, defaultdict
import os
from pathlib import Path

from openpyxl import load_workbook


CP_ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = Path(
    os.environ.get(
        "SCHEDULE_WORKBOOK",
        CP_ROOT / "ICPC_Solo_Training_Schedule_2026_2027.xlsx",
    )
)


class SoloTrainingScheduleTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        workbook = load_workbook(WORKBOOK, data_only=False, read_only=True)
        cls.sheet_names = workbook.sheetnames
        sheet = workbook["Solo Timeline"]
        cls.headers = [cell.value for cell in sheet[1]]
        cls.index = {name: idx for idx, name in enumerate(cls.headers)}
        cls.rows = [
            row
            for row in sheet.iter_rows(min_row=2, values_only=True)
            if any(value is not None for value in row)
        ]

    def value(self, row, column):
        return row[self.index[column]]

    def test_workbook_shape_and_formula_health(self):
        self.assertEqual(
            self.sheet_names,
            [
                "Dashboard",
                "Profile Basis",
                "Solo Timeline",
                "Weekly Brief",
                "All Links",
                "Sources",
                "Lists",
            ],
        )
        bad_formulas = []
        for row in self.rows:
            for value in row:
                if isinstance(value, str) and ("#REF!" in value or "#VALUE!" in value):
                    bad_formulas.append(value)
        self.assertEqual(bad_formulas, [])

    def test_weekly_loads_respect_caps_and_remain_intensive(self):
        hours = defaultdict(float)
        for row in self.rows:
            hours[int(self.value(row, "Week"))] += float(
                self.value(row, "Expected hours") or 0
            )

        for week in range(1, 9):
            self.assertGreaterEqual(hours[week], 40.0, f"W{week} is too light")
            self.assertLessEqual(hours[week], 45.0, f"W{week} exceeds camp cap")
        for week in range(9, 20):
            self.assertGreaterEqual(hours[week], 32.0, f"W{week} is too light")
            self.assertLessEqual(hours[week], 35.0, f"W{week} exceeds build cap")
        self.assertEqual(hours[20], 22.0)
        self.assertEqual(hours[21], 10.0)

    def test_pre_regional_judged_mix(self):
        judged = [
            row
            for row in self.rows
            if int(self.value(row, "Week")) <= 19
            and self.value(row, "Work type") in {"Learn", "Hard"}
            and self.value(row, "Source type") != "Self practice"
        ]
        sources = Counter(self.value(row, "Source type") for row in judged)

        self.assertGreaterEqual(len(judged), 70)
        self.assertLessEqual(sources["CSES"], 10)
        modern = sum(
            sources[source]
            for source in (
                "QOJ / Universal Cup",
                "Codeforces Gym",
                "OI / oj.uz",
                "Kattis ICPC",
            )
        )
        self.assertGreaterEqual(modern, 15)
        self.assertGreaterEqual(sources["OI / oj.uz"], 1)
        self.assertGreaterEqual(sources["ICPC / VNOI"], 18)

        topic_counts = Counter(
            self.value(row, "Topic")
            for row in judged
            if str(self.value(row, "Topic")).startswith(tuple(f"0{i}" for i in range(9)))
        )
        self.assertEqual(len(topic_counts), 9)
        self.assertGreaterEqual(min(topic_counts.values()), 7)

    def test_learn_support_is_concrete_and_linked(self):
        support = [
            row
            for row in self.rows
            if int(self.value(row, "Week")) <= 19
            and self.value(row, "Work type") == "Learn"
            and self.value(row, "Source type") == "Self practice"
        ]
        total_hours = sum(float(self.value(row, "Expected hours") or 0) for row in support)

        self.assertGreaterEqual(len(support), 15)
        self.assertLessEqual(len(support), 25)
        self.assertGreaterEqual(total_hours, 55.0)
        self.assertLessEqual(total_hours, 80.0)
        self.assertEqual(Counter(int(self.value(row, "Week")) for row in support), Counter(range(1, 20)))
        for row in support:
            name = str(self.value(row, "Task name"))
            self.assertNotIn("Learn support: read + implement core technique", name)
            self.assertTrue(self.value(row, "Editorial/blog"), name)
            self.assertTrue(self.value(row, "Why this problem"), name)

    def test_judged_tasks_have_no_duplicates(self):
        names = [
            self.value(row, "Task name")
            for row in self.rows
            if self.value(row, "Work type") in {"Learn", "Hard"}
            and self.value(row, "Source type") != "Self practice"
        ]
        duplicates = [name for name, count in Counter(names).items() if count > 1]
        self.assertEqual(duplicates, [])

    def test_late_regional_virtuals_are_clean_sets(self):
        expected = {
            16: "Chengdu Regional 2025 (virtual 5h)",
            17: "Nanjing Regional 2025 (virtual 5h)",
            18: "Shenyang Regional 2025 (virtual 5h)",
            19: "Shanghai Regional 2025 (virtual 5h)",
        }
        actual = {
            int(self.value(row, "Week")): self.value(row, "Task name")
            for row in self.rows
            if self.value(row, "Work type") == "Contest"
            and 16 <= int(self.value(row, "Week")) <= 19
        }
        self.assertEqual(actual, expected)

        contest_sources = {
            16: "2025 ICPC Asia Chengdu Regional",
            17: "2025 ICPC Asia Nanjing Regional",
            18: "2025 ICPC Asia Shenyang Regional",
            19: "2025 ICPC Asia Shanghai Regional",
        }
        for contest_week, contest_source in contest_sources.items():
            overlap = [
                self.value(row, "Task name")
                for row in self.rows
                if int(self.value(row, "Week")) < contest_week
                and self.value(row, "Work type") in {"Learn", "Hard"}
                and self.value(row, "Contest/source") == contest_source
            ]
            self.assertEqual(overlap, [], contest_source)


if __name__ == "__main__":
    unittest.main(verbosity=2)
